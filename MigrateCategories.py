#!/usr/bin/python

import boto3
import botocore
import logging
import openpyxl
import csv
import psycopg2
import os
import sys
from datetime import datetime
from GetDBConfigParam import GetDBConfigParam 
from GetS3Session import GetS3Session 


def process_Categories_data(DBParams, s3_bucket, local_file_name):

    conn = None
    file_total_rows = 0
    db_inserted_rows = 0
    db_updated_rows = 0
    db_deleted_rows = 0
    file_nochange_rows = 0

    try:
        # DB connection parameters
        params = DBParams

        # Connect to the PostgreSQL database server
        logging.info('Connecting to the PostgreSQL database...')
        #conn = psycopg2.connect(**params)
        conn = psycopg2.connect(database=params.path[1:],
                                user=params.username,
                                password=params.password,
                                host=params.hostname,
                                port=params.port)

        # create a cursor
        cur = conn.cursor()

        # display the PostgreSQL database server version, name & user
        # cur.execute('SELECT version()')
        # db_version = cur.fetchone()
        # logging.info('PostgreSQL database version: ' + str(db_version))

        cur.execute('SELECT current_database()')
        db_name = cur.fetchone()
        logging.info('PostgreSQL database name: ' + str(db_name))

        cur.execute('SELECT current_user')
        db_user = cur.fetchone()
        logging.info('PostgreSQL database user name: ' + str(db_user))

        logging.info('Database is successfully connected...')

        # Set current date, time variables
        Now = datetime.now()

        # Set file location
        logging.info('Workbook, Sheet Name: ' + str(local_file_name) + ', Categories')
        logging.info('Workbook Location: ' + str(os.getcwd())) 


        # Set Sheet for read
        book = openpyxl.load_workbook(filename = local_file_name)
        sheet = book['Categories']
        file_total_rows = sheet.max_row             #Reconciliation
        logging.info('Total Columns: ' + str(sheet.max_column) + ', Total Rows: ' + str(sheet.max_row))

        # Instructions and Header Processing 
        # CategoriesKey = sheet.cell(row=2, column=1).value
        # Delete = sheet.cell(row=2, column=2).value
        # Name = sheet.cell(row=2, column=3).value
        # Order = sheet.cell(row=2, column=4).value
        # ParentKey = sheet.cell(row=2, column=5).value
        # logging.info('Header: ' + str(CategoriesKey) + ', ' + str(Delete) + ', ' + str(Name) + ', ' + str(Order) + ', ' + str(ParentKey))
        logging.info('Instructions and Header lines, no change required')
        file_nochange_rows = file_nochange_rows + 2         #Reconciliation Hearder row

        # Read all lines after header and process them
        for r in range(3, sheet.max_row+1):
            CategoriesKey = sheet.cell(row=r, column=1).value
            Delete = sheet.cell(row=r, column=2).value
            Name = sheet.cell(row=r, column=3).value
            Order = sheet.cell(row=r, column=4).value
            ParentKey = sheet.cell(row=r, column=5).value

            if (Delete not in ['Yes', 'YES', 'yes']):
                Status = 'Active'
            else:
                Status = 'Inactive'

            sql_getParentID = """ SELECT id
                            FROM public.categories
                            WHERE key = %s AND status = %s;"""
            cur.execute(sql_getParentID, (ParentKey, 'Active'))

            dbParentID = cur.fetchone()
            
            if dbParentID is None:
                ParentID = None
            else:
                ParentID = dbParentID[0]

            # fileValues = (CategoriesKey, Delete, Status, Name, Order, ParentKey, ParentID)
            # logging.info('File data: ' + str(fileValues))
                
            # For each record read, check if it exist in the table
            sql_exist = """ SELECT name,
                                "order",
                                "parentId",
                                "createdAt",
                                "updatedAt",
                                status,
                                key,
                                id
                            FROM public.categories
                            WHERE key = %s"""

            cur.execute(sql_exist, (CategoriesKey,))
            row_exist = cur.fetchone()

            #if row_exist is not None:
            #    logging.info('DB data: ' + str(row_exist[0:]))
            #    logging.info('DB ID: ' + str(row_exist[7]))
            #else:
            #    logging.info('No DB Data')
                
            # If row does not exist = Insert the record    
            if row_exist is None:
                #logging.info('Data insert')

                sql_insert = """ INSERT INTO public.categories(name,
                                                    "order",
                                                    "parentId",
                                                    "createdAt",
                                                    "updatedAt",
                                                    status,
                                                    key)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                RETURNING id;"""

                cur.execute(sql_insert, (Name, Order, ParentID, Now, Now, Status, CategoriesKey))
                inserted_id = cur.fetchone()[0]
                inserted_rows = cur.rowcount
                db_inserted_rows = db_inserted_rows + inserted_rows             #Reconciliation
                conn.commit()
                logging.info(str(inserted_rows) + 'Record/s inserted with ID, Key: ' + str(inserted_id) + ', ' + str(CategoriesKey))

            # If row exist = Update record if any value or Logical delete is changed
            elif ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Inactive') or                #Activated Inactive record
                  (Delete not in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Active' and                  #Updated Active record     
                    (Name != row_exist[0] or
                    Order!= row_exist[1] or
                    ParentID != row_exist[2])) or
                  (Delete in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Active')):                       #Inactivated Active record
                
                if (Delete not in ['Yes', 'YES', 'yes']):
                    Status = 'Active'
                else:
                    Status = 'Inactive'
                    
                sql_update = """ UPDATE public.categories
                                    SET name = %s,
                                        "order" = %s, 
                                        "parentId" = %s, 
                                        "updatedAt" = %s,
                                        status = %s
                                    WHERE key = %s;"""
                cur.execute(sql_update, (Name, Order, ParentID, Now, Status, CategoriesKey))
                
                if ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Inactive') or
                    (Delete not in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Active')):
                    #logging.info('Data update')
                    updated_rows = cur.rowcount
                    conn.commit()
                    db_updated_rows = db_updated_rows + updated_rows             #Reconciliation
                    logging.info(str(updated_rows) + 'Record/s updated for ID, Key: ' + str(row_exist[7]) + ', ' + str(CategoriesKey))
                else:
                    #logging.info('Data delete')
                    deleted_rows = cur.rowcount
                    conn.commit()
                    db_deleted_rows = db_deleted_rows + deleted_rows             #Reconciliation
                    logging.info(str(deleted_rows) + 'Record/s logically deleted for ID, Key: ' + str(row_exist[7]) + ', ' + str(CategoriesKey))

            # If row exist = No action if Delete is "Yes" and record is inactive
            elif (Delete in ['Yes', 'YES', 'yes'] and row_exist[5] == 'Inactive'):
                logging.info('file/DB data Inactive, no change required for ID, Key:' + str(row_exist[7]) + ', ' + str(CategoriesKey))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

            # If row exist = No action if no file data is changed
            else:
                logging.info('file data is not changed for ID, Key:' + str(row_exist[7]) + ', ' + str(CategoriesKey))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

        # close a cursor
        cur.close()
        
    except (Exception, psycopg2.DatabaseError) as error:
        logging.error(error)

    finally:
        if conn is not None:
            conn.close()
            logging.info('Database connection closed.')

        reconciliation_data = ['categories', file_total_rows, db_inserted_rows, db_updated_rows, db_deleted_rows, file_nochange_rows]
        logging.info('Recon Data: ' + str(reconciliation_data))
        
    return reconciliation_data

if __name__ == '__main__':

    # Get time
    t = (str(datetime.now().year), str(datetime.now().month), str(datetime.now().day),
         str(datetime.now().hour), str(datetime.now().minute), str(datetime.now().second))

    # Get Log file defined
    # Log level "DEBUG" (logging.DEBUG) generate many debug log messages from Boto3
    LogFilename = "-".join(t) + '_' + r'logfile.log'
    logging.basicConfig(level=logging.INFO, handlers=[logging.FileHandler(LogFilename, 'a+', 'utf-8')],
                            format="%(asctime)-15s - %(levelname)-8s %(message)s")
    logging.info('Log file: ' + str(LogFilename))

    # Recon file
    ReconFilename = "-".join(t) + '_' + r'reconfile.csv'
    ReconFile = open(ReconFilename, "w")
    ReconFilewriter = csv.writer(ReconFile, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
    logging.info('Recon file: ' + str(ReconFilename))

    # Recon file header
    ReconHeader = ['FileName' , 'File Total Rows' , 'DB Inserted Rows' , 'DB Updated Rows' , 'DB Deleted Rows' , 'File Nochange Rows']
    ReconFilewriter.writerow(ReconHeader)

    # read connection parameters
    params = GetDBConfigParam()
    DBParams = params

    # get S3 session established
    session = GetS3Session()

    # get the S3 bucket
    s3_resource = session.resource('s3')
    bucket_name = 'rmit60.ks'
    s3_bucket = s3_resource.Bucket(bucket_name)

    # check for the file and process it
    POIsFilename = 'Wayfinding Locations.xlsx'
    key_file_name = POIsFilename
    POIsLocalFilename = "-".join(t) + '_' + POIsFilename
    local_file_name = POIsLocalFilename
    try:
        s3_bucket.download_file(key_file_name, local_file_name)
        logging.info('================================================================================')
        logging.info(str(key_file_name) + ' available for local processing as ' + str(local_file_name) + '.')
        archive_file_name = 'archive/' + local_file_name
        s3_bucket.upload_file(local_file_name, archive_file_name)
        logging.info(str(local_file_name) + ' is transferred to S3 bucket.')
        logging.info('================================================================================')
        logging.info('Processing Categories...')

        Categories_reconciliation_data = process_Categories_data(params, s3_bucket, local_file_name)
        ReconFilewriter.writerow([Categories_reconciliation_data[0], Categories_reconciliation_data[1], Categories_reconciliation_data[2],
                                 Categories_reconciliation_data[3], Categories_reconciliation_data[4], Categories_reconciliation_data[5]])

        # Delete source files as they are processed and copied to S3 additing timestamp
        logging.info(str(key_file_name) + ' will be deleted from S3 as ' + str(local_file_name) + ' is added to S3.')
        # Delete is not required for single file processing
        #s3_bucket.Object(key_file_name).delete()

    except botocore.exceptions.ClientError as error:
            if error.response['Error']['Code'] == "404":
                logging.info('================================================================================')                
                logging.info(str(key_file_name) + ' not available for processing.')
                logging.info('================================================================================')
                ReconFilewriter.writerow(['categories', 0, 0, 0, 0, 0])
            else:
                logging.error(error)


    # Recon file close and move to S3 
    ReconFile.close()
    local_file_name = ReconFilename
    archive_file_name = 'archive/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
    logging.info(str(ReconFilename) + ' is transferred to S3 bucket.')

    # Log file close and move to S3
    logging.info(str(LogFilename) + ' will be closed and then transferred to S3 bucket.')
    logging.info('End of the Job')
    logging.shutdown()
    local_file_name = LogFilename
    archive_file_name = 'archive/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
