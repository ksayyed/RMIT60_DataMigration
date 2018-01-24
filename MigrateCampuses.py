#!/usr/bin/python

import boto3
import botocore
import logging
import openpyxl
import csv
import psycopg2
import os
import pytz
import sys
from datetime import datetime
from GetDBConfigParam import GetDBConfigParam
from GetS3Session import GetS3Session

def process_Campuses_data(DBParams, s3_bucket, local_file_name):
    
    conn = None
    file_total_rows = 0
    db_inserted_rows = 0
    db_updated_rows = 0
    db_deleted_rows = 0
    file_nochange_rows = 0

    try:
        # DB connection parameters
        params = DBParams
        
        # Connect to the PostgreSQL database server
        logging.info('Connecting to the PostgreSQL database...')
        #conn = psycopg2.connect(**params)
        conn = psycopg2.connect(database=params.path[1:],
                                user=params.username,
                                password=params.password,
                                host=params.hostname,
                                port=params.port)

        # create a cursor
        cur = conn.cursor()

        # display the PostgreSQL database server version, name & user
        # cur.execute('SELECT version()')
        # db_version = cur.fetchone()
        # logging.info('PostgreSQL database version: ' + str(db_version))

        cur.execute('SELECT current_database()')
        db_name = cur.fetchone()
        logging.info('PostgreSQL database name: ' + str(db_name))

        cur.execute('SELECT current_user')
        db_user = cur.fetchone()
        logging.info('PostgreSQL database user name: ' + str(db_user))

        logging.info('Database is successfully connected...')

        # Set current date, time variables
        Now = datetime.now()

        # Set file location
        logging.info('Workbook, Sheet Name: ' + str(local_file_name) + ', Campuses')
        logging.info('Workbook Location: ' + str(os.getcwd())) 

        # Set Sheet for read
        book = openpyxl.load_workbook(filename = local_file_name)
        sheet = book['Campuses']
        file_total_rows = sheet.max_row             #Reconciliation
        logging.info('Total Columns: ' + str(sheet.max_column) + ', Total Rows: ' + str(sheet.max_row))

        # Instructions and Header Processing 
        # Id = sheet.cell(row=2, column=1).value
        # Delete = sheet.cell(row=2, column=2).value
        # Name = sheet.cell(row=2, column=3).value
        # CenterLatitude = sheet.cell(row=2, column=4).value
        # CenterLongitude = sheet.cell(row=2, column=5).value
        # ZoomLevel = sheet.cell(row=2, column=6).value
        # logging.info('Header: ' + str(Id) + ', ' + str(Delete) + ', ' + str(Name) + ', ' + str(CenterLatitude) + ', ' + str(CenterLongitude) + ', ' + str(ZoomLevel))
        logging.info('Instructions and Header lines, no change required')
        file_nochange_rows = file_nochange_rows + 2         #Reconciliation Hearder row

        # Read all lines after header and process them
        for r in range(3, sheet.max_row+1):
            Id = sheet.cell(row=r, column=1).value
            Delete = sheet.cell(row=r, column=2).value
            Name = sheet.cell(row=r, column=3).value
            CenterLatitude = sheet.cell(row=r, column=4).value
            CenterLongitude = sheet.cell(row=r, column=5).value
            ZoomLevel = sheet.cell(row=r, column=6).value

            if (Delete not in ['Yes', 'YES', 'yes']):
                Status = 'Active'
            else:
                Status = 'Inactive'

            # fileValues = (Id, Delete, Name, CenterLatitude, CenterLongitude, ZoomLevel)
            # logging.info('File data: ' + str(fileValues))

            # For each record read, check if it exist in the table
            sql_exist = """ SELECT name, 
                                latitude, 
                                longitude, 
                                "zoomLevel", 
                                "createdAt", 
                                "updatedAt", 
                                status,
                                id
                            FROM public.campuses
                            WHERE id = %s;"""
            cur.execute(sql_exist, (Id,))
            row_exist = cur.fetchone()

            #if row_exist is not None:
            #    logging.info('DB data: ' + str(row_exist[0:7]))
            #    logging.info('DB ID: ' + str(row_exist[7]))
            #else:
            #    logging.info('No DB Data')
                
            # If row does not exist = Insert the record    
            if row_exist is None:
                #logging.info('Data inset')

                sql_insert = """ INSERT INTO public.campuses
                                                        (name, 
                                                        latitude, 
                                                        longitude, 
                                                        "zoomLevel",
                                                        "createdAt",
                                                        "updatedAt",
                                                        status,
                                                        id)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s);"""

                cur.execute(sql_insert, (Name, CenterLatitude, CenterLongitude, ZoomLevel, Now, Now, Status, Id))
                #inserted_id = cur.fetchone()[0]
                inserted_id = Id
                inserted_rows = cur.rowcount
                db_inserted_rows = db_inserted_rows + inserted_rows             #Reconciliation
                conn.commit()
                logging.info(str(inserted_rows) + 'Record/s inserted with ID, Name: ' + str(inserted_id) + ', ' + str(Name))

            # If row exist = Update record if any value or Logical delete is changed
            elif ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Inactive') or                #Activated Inactive record
                  (Delete not in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Active' and                  #Updated Active record     
                    (str(Name) != str(row_exist[0]) or
                    #str(CenterLatitude) != str(row_exist[1]) or
                    ((CenterLatitude is not None and row_exist[1] is None) or 
                     (CenterLatitude is None and row_exist[1] is not None) or 
                     (CenterLatitude is not None and row_exist[1] is not None and str(CenterLatitude) != str(row_exist[1]))) or
                    #str(CenterLongitude) != str(row_exist[2]) or
                    ((CenterLongitude is not None and row_exist[2] is None) or 
                     (CenterLongitude is None and row_exist[2] is not None) or 
                     (CenterLongitude is not None and row_exist[2] is not None and str(CenterLongitude) != str(row_exist[2]))) or
                    #str(ZoomLevel) != str(row_exist[3])
                    ((ZoomLevel is not None and row_exist[3] is None) or 
                     (ZoomLevel is None and row_exist[3] is not None) or 
                     (ZoomLevel is not None and row_exist[3] is not None and str(ZoomLevel) != str(row_exist[3]))))) or                  
                  (Delete in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Active')):                       #Inactivated Active record
                
                sql_update = """ UPDATE public.campuses
                                    SET name = %s,
                                        latitude = %s, 
                                        longitude = %s, 
                                        "zoomLevel" = %s,
                                        "updatedAt" = %s,
                                        status = %s
                                    WHERE id = %s;"""
                cur.execute(sql_update, (Name, CenterLatitude, CenterLongitude, ZoomLevel, Now, Status, Id))
                
                if ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Inactive') or
                    (Delete not in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Active')):
                    #logging.info('Data update')
                    updated_rows = cur.rowcount
                    conn.commit()
                    db_updated_rows = db_updated_rows + updated_rows             #Reconciliation
                    logging.info(str(updated_rows) + 'Record/s updated for ID, Name: ' + str(row_exist[7]) + ', ' + str(row_exist[0]))
                else:
                    #logging.info('Data delete')
                    deleted_rows = cur.rowcount
                    conn.commit()
                    db_deleted_rows = db_deleted_rows + deleted_rows             #Reconciliation
                    logging.info(str(deleted_rows), 'Record/s logically deleted for ID, Name: ' + str(row_exist[7]) + ', ' + str(row_exist[0]))

            # If row exist = No action if Delete is "Yes" and record is inactive
            elif (Delete in ['Yes', 'YES', 'yes'] and row_exist[6] == 'Inactive'):
                logging.info('file/DB data Inactive, no change required for ID, Name: ' + str(row_exist[7]) + ', ' + str(row_exist[0]))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

            # If row exist = No action if no file data is changed
            else:
                logging.info('file data is not changed for ID, Name: ' + str(row_exist[7]) + ', ' + str(row_exist[0]))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

        # close a cursor
        cur.close()
        
    except (Exception, psycopg2.DatabaseError) as error:
        logging.error(error)

    finally:
        if conn is not None:
            conn.close()
            logging.info('Database connection closed.')

        reconciliation_data = ['campuses', file_total_rows, db_inserted_rows, db_updated_rows, db_deleted_rows, file_nochange_rows]
        logging.info('Recon Data: ' + str(reconciliation_data))
        
    return reconciliation_data


if __name__ == '__main__':

    # Get time
    au_tz = str(os.environ["TZ"])
    now = datetime.now(pytz.timezone(au_tz))
    t = (str(now.year), str(now.month), str(now.day), str(now.hour), str(now.minute), str(now.second))

    # Get Log file defined
    # Log level "DEBUG" (logging.DEBUG) generate many debug log messages from Boto3
    LogFilename = "-".join(t) + '_' + r'logfile.log'
    logging.basicConfig(level=logging.INFO, handlers=[logging.FileHandler(LogFilename, 'a+', 'utf-8')],
                            format="%(asctime)-15s - %(levelname)-8s %(message)s")
    logging.info('Log file: ' + str(LogFilename))

    # Recon file
    ReconFilename = "-".join(t) + '_' + r'reconfile.csv'
    ReconFile = open(ReconFilename, "w")
    ReconFilewriter = csv.writer(ReconFile, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
    logging.info('Recon file: ' + str(ReconFilename))

    # Recon file header
    ReconHeader = ['FileName' , 'File Total Rows' , 'DB Inserted Rows' , 'DB Updated Rows' , 'DB Deleted Rows' , 'File Nochange Rows']
    ReconFilewriter.writerow(ReconHeader)

    # read connection parameters
    params = GetDBConfigParam()
    DBParams = params

    # get S3 session established
    session = GetS3Session()

    # get the S3 bucket
    s3_resource = session.resource('s3')
    bucket_name = str(os.environ["AWS_BUCKET_NAME"])
    s3_bucket = s3_resource.Bucket(bucket_name)
    logging.info('S3 bucket name: ' + str(bucket_name))

    # check for the file and process it
    POIsFilename = 'Wayfinding Locations.xlsx'
    key_file_name = POIsFilename
    POIsLocalFilename = "-".join(t) + '_' + POIsFilename
    local_file_name = POIsLocalFilename
    try:
        s3_bucket.download_file(key_file_name, local_file_name)
        logging.info('================================================================================')
        logging.info(str(key_file_name) + ' available for local processing as ' + str(local_file_name) + '.')
        archive_file_name = 'archive/' + "-".join(t) + '/' + local_file_name
        s3_bucket.upload_file(local_file_name, archive_file_name)
        logging.info(str(local_file_name) + ' is transferred to S3 bucket.')
        logging.info('================================================================================')
        logging.info('Processing Campuses...')

        Campuses_reconciliation_data = process_Campuses_data(params, s3_bucket, local_file_name)
        ReconFilewriter.writerow([Campuses_reconciliation_data[0], Campuses_reconciliation_data[1], Campuses_reconciliation_data[2],
                                 Campuses_reconciliation_data[3], Campuses_reconciliation_data[4], Campuses_reconciliation_data[5]])

        # Delete is not required for single file processing
        # Delete source files as they are processed and copied to S3 additing timestamp
        #logging.info(str(key_file_name) + ' will be deleted from S3 as ' + str(local_file_name) + ' is added to S3.')
        #s3_bucket.Object(key_file_name).delete()

    except botocore.exceptions.ClientError as error:
            if error.response['Error']['Code'] == "404":
                logging.info('================================================================================')                
                logging.info(str(key_file_name) + ' not available for processing.')
                logging.info('================================================================================')
                ReconFilewriter.writerow(['campuses', 0, 0, 0, 0, 0])
            else:
                logging.error(error)


    # Recon file close and move to S3 
    ReconFile.close()
    local_file_name = ReconFilename
    archive_file_name = 'archive/' + "-".join(t) + '/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
    logging.info(str(ReconFilename) + ' is transferred to S3 bucket.')

    # Log file close and move to S3
    logging.info(str(LogFilename) + ' will be closed and then transferred to S3 bucket.')
    logging.info('End of the Job')
    logging.shutdown()
    local_file_name = LogFilename
    archive_file_name = 'archive/' + "-".join(t) + '/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
