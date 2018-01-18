#!/usr/bin/python

import boto3
import botocore
import logging
import openpyxl
import csv
import psycopg2
import os
import sys
from datetime import datetime
from GetDBConfigParam import GetDBConfigParam
from GetS3Session import GetS3Session

def process_Tasks_data(DBParams, s3_bucket, local_file_name):

    conn = None
    file_total_rows = 0
    db_inserted_rows = 0
    db_updated_rows = 0
    db_deleted_rows = 0
    file_nochange_rows = 0

    try:
        # DB connection parameters
        params = DBParams

        # Connect to the PostgreSQL database server
        logging.info('Connecting to the PostgreSQL database...')
        #conn = psycopg2.connect(**params)
        conn = psycopg2.connect(database=params.path[1:],
                                user=params.username,
                                password=params.password,
                                host=params.hostname,
                                port=params.port)

        # create a cursor
        cur = conn.cursor()

        # display the PostgreSQL database server version, name & user
        # cur.execute('SELECT version()')
        # db_version = cur.fetchone()
        # logging.info('PostgreSQL database version: ' + str(db_version))

        cur.execute('SELECT current_database()')
        db_name = cur.fetchone()
        logging.info('PostgreSQL database name: ' + str(db_name))

        cur.execute('SELECT current_user')
        db_user = cur.fetchone()
        logging.info('PostgreSQL database user name: ' + str(db_user))

        logging.info('Database is successfully connected...')

        # Set current date, time variables
        Now = datetime.now()

        # Set file location
        logging.info('Workbook, Sheet Name: ' + str(local_file_name) + ', Tasks')
        logging.info('Workbook Location: ' + str(os.getcwd())) 

        # Set Sheet for read
        book = openpyxl.load_workbook(filename = local_file_name)
        sheet = book['Tasks']
        file_total_rows = sheet.max_row             #Reconciliation
        logging.info('Total Columns: ' + str(sheet.max_column) + ', Total Rows: ' + str(sheet.max_row))

        # Instructions and Header Processing 
        # TaskKey = sheet.cell(row=2, column=1).value
        # Delete = sheet.cell(row=2, column=2).value
        # Question = sheet.cell(row=2, column=3).value
        # Title = sheet.cell(row=2, column=4).value
        # Description = sheet.cell(row=2, column=5).value
        # CTA = sheet.cell(row=2, column=6).value
        # URL = sheet.cell(row=2, column=7).value
        # StartDate = sheet.cell(row=2, column=8).value
        # EndDate = sheet.cell(row=2, column=9).value
        # PhaseKey = sheet.cell(row=2, column=10).value
        # logging.info('Header: ' + str(TaskKey) + ', ' + str(Delete) + ', ' + str(Title) + ', ' + str(Question) + ', ' + str(Title) + ', ' + str(Description) + ', '
        #               + str(CTA) + ', ' + str(URL) + ', ' + str(StartDate) + ', ' + str(EndDate) + ', ' + str(PhaseKey))
        logging.info('Instructions and Header lines, no change required')
        file_nochange_rows = file_nochange_rows + 2         #Reconciliation Hearder row

        # Read all lines after header and process them
        for r in range(3, sheet.max_row+1):
            TaskKey = sheet.cell(row=r, column=1).value
            Delete = sheet.cell(row=r, column=2).value
            Question = sheet.cell(row=r, column=3).value
            Title = sheet.cell(row=r, column=4).value
            Description = sheet.cell(row=r, column=5).value
            CTA = sheet.cell(row=r, column=6).value
            URL = sheet.cell(row=r, column=7).value
            Start = sheet.cell(row=r, column=8).value
            End = sheet.cell(row=r, column=9).value
            PhaseKey = sheet.cell(row=r, column=10).value

            if Start is None:
                StartDate = Start
            else:
                StartDate = Start.strftime('%Y-%m-%d')

            if End is None:
                EndDate = End
            else:
                EndDate = End.strftime('%Y-%m-%d')

            if (Delete not in ['Yes', 'YES', 'yes']):
                Status = 'Active'
            else:
                Status = 'Inactive'

            sql_getPhaseID = """ SELECT id
                            FROM public.phases
                            WHERE key = %s AND status = %s;"""
            cur.execute(sql_getPhaseID, (PhaseKey, 'Active'))

            dbphaseid = cur.fetchone()
            
            if dbphaseid is None:
                PhaseId = None
            else:
                PhaseId = dbphaseid[0]

            # fileValues = (TaskKey, Delete, Status, Question, Title, Description, CTA, URL, StartDate, EndDate, PhaseKey, PhaseId)
            # logging.info('File data: ' + str(fileValues))
                
            # For each record read, check if it exist in the table
            sql_exist = """ SELECT title, 
                                question, 
                                description, 
                                cta, 
                                url, 
                                "startDate", 
                                "endDate", 
                                "phaseId",
                                status,
                                key, 
                                id
                            FROM public.tasks
                            WHERE key = %s"""

            cur.execute(sql_exist, (TaskKey,))
            row_exist = cur.fetchone()

            dbstart =None
            dbstartdate = None
            if row_exist is not None:
                dbstart = row_exist[5]
                if dbstart is not None:
                    dbstartdate = dbstart.strftime('%Y-%m-%d')

            dbend = None
            dbenddate = None
            if row_exist is not None:
                dbend = row_exist[6]
                if dbend is not None:
                    dbenddate = dbend.strftime('%Y-%m-%d')

            #if row_exist is not None:
            #   logging.info('DB data: ' + str(row_exist[0:]))
            #   logging.info('DB ID: ' + str(row_exist[10]))
            #else:
            #   logging.info('No DB Data')

            # If row does not exist = Insert the record    
            if row_exist is None:
                #logging.info('Data inset')

                sql_insert = """ INSERT INTO public.tasks(title,
                                                        question, 
                                                        description, 
                                                        cta, 
                                                        url, 
                                                        "startDate", 
                                                        "endDate", 
                                                        "phaseId", 
                                                        "createdAt",
                                                        "updatedAt",
                                                        status,
                                                        key)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                RETURNING id;"""

                cur.execute(sql_insert, (Title, Question, Description, CTA, URL, StartDate, EndDate, PhaseId, Now, Now, Status, TaskKey))
                inserted_id = cur.fetchone()[0]
                inserted_rows = cur.rowcount
                db_inserted_rows = db_inserted_rows + inserted_rows             #Reconciliation
                conn.commit()
                logging.info(str(inserted_rows) + 'Record/s inserted with ID, Key: ' + str(inserted_id) + ', ' + str(TaskKey))

            # If row exist = Update record if any value or Logical delete is changed
            elif ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Inactive') or                #Activated Inactive record
                  (Delete not in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Active' and                  #Updated Active record     
                    (Title != row_exist[0] or
                    Question!= row_exist[1] or
                    Description != row_exist[2] or 
                    CTA != row_exist[3] or
                    URL != row_exist[4] or
                    StartDate != dbstartdate or
                    EndDate != dbenddate or
                    PhaseId != row_exist[7] )) or
                  (Delete in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Active')):                       #Inactivated Active record
                
                if (Delete not in ['Yes', 'YES', 'yes']):
                    Status = 'Active'
                else:
                    Status = 'Inactive'
                    
                sql_update = """ UPDATE public.tasks
                                    SET title = %s,
                                        question = %s, 
                                        description = %s, 
                                        cta = %s,
                                        url = %s, 
                                        "startDate" = %s, 
                                        "endDate" = %s, 
                                        "phaseId" = %s, 
                                        "updatedAt" = %s,
                                        status = %s
                                    WHERE key = %s;"""
                cur.execute(sql_update, (Title, Question, Description, CTA, URL, StartDate, EndDate, PhaseId, Now, Status, TaskKey))
                
                if ((Delete not in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Inactive') or
                    (Delete not in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Active')):
                    #logging.info('Data update')
                    updated_rows = cur.rowcount
                    conn.commit()
                    db_updated_rows = db_updated_rows + updated_rows             #Reconciliation
                    logging.info(str(updated_rows) + 'Record/s updated for ID, Key: ' + str(row_exist[10]) + ', ' + str(TaskKey))
                else:
                    #logging.info('Data delete')
                    deleted_rows = cur.rowcount
                    conn.commit()
                    db_deleted_rows = db_deleted_rows + deleted_rows             #Reconciliation
                    logging.info(str(deleted_rows) + 'Record/s logically deleted for ID, Key: ' + str(row_exist[10]) + ', ' + str(TaskKey))

            # If row exist = No action if Delete is "Yes" and record is inactive
            elif (Delete in ['Yes', 'YES', 'yes'] and row_exist[8] == 'Inactive'):
                logging.info('file/DB data Inactive, no change required for ID, Key:' + str(row_exist[10]) + ', ' + str(TaskKey))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

            # If row exist = No action if no file data is changed
            else:
                logging.info('file data is not changed for ID, Key:' + str(row_exist[10]) + ', ' + str(TaskKey))
                nochange_rows = 1
                file_nochange_rows = file_nochange_rows + nochange_rows             #Reconciliation

        # close a cursor
        cur.close()
        
    except (Exception, psycopg2.DatabaseError) as error:
        logging.error(error)

    finally:
        if conn is not None:
            conn.close()
            logging.info('Database connection closed.')

        reconciliation_data = ['tasks', file_total_rows, db_inserted_rows, db_updated_rows, db_deleted_rows, file_nochange_rows]
        logging.info('Recon Data: ' + str(reconciliation_data))
        
    return reconciliation_data

if __name__ == '__main__':

    # Get time
    t = (str(datetime.now().year), str(datetime.now().month), str(datetime.now().day),
         str(datetime.now().hour), str(datetime.now().minute), str(datetime.now().second))

    # Get Log file defined
    # Log level "DEBUG" (logging.DEBUG) generate many debug log messages from Boto3
    LogFilename = "-".join(t) + '_' + r'logfile.log'
    logging.basicConfig(level=logging.INFO, handlers=[logging.FileHandler(LogFilename, 'a+', 'utf-8')],
                            format="%(asctime)-15s - %(levelname)-8s %(message)s")
    logging.info('Log file: ' + str(LogFilename))

    # Recon file
    ReconFilename = "-".join(t) + '_' + r'reconfile.csv'
    ReconFile = open(ReconFilename, "w")
    ReconFilewriter = csv.writer(ReconFile, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
    logging.info('Recon file: ' + str(ReconFilename))

    # Recon file header
    ReconHeader = ['FileName' , 'File Total Rows' , 'DB Inserted Rows' , 'DB Updated Rows' , 'DB Deleted Rows' , 'File Nochange Rows']
    ReconFilewriter.writerow(ReconHeader)

    # read connection parameters
    params = GetDBConfigParam()
    DBParams = params

    # get S3 session established
    session = GetS3Session()

    # get the S3 bucket
    s3_resource = session.resource('s3')
    bucket_name = 'rmit60.ks'
    s3_bucket = s3_resource.Bucket(bucket_name)

    # check for the file and process it
    TasksFilename = 'Onboarding Tasks.xlsx'
    key_file_name = TasksFilename
    TasksLocalFilename = "-".join(t) + '_' + TasksFilename 
    local_file_name = TasksLocalFilename
    try:
        s3_bucket.download_file(key_file_name, local_file_name)
        logging.info('================================================================================')
        logging.info(str(key_file_name) + ' available for local processing as ' + str(local_file_name) + '.')
        archive_file_name = 'archive/' + local_file_name
        s3_bucket.upload_file(local_file_name, archive_file_name)
        logging.info(str(local_file_name) + ' is transferred to S3 bucket.')
        logging.info('================================================================================')
        logging.info('Processing Tasks...')

        Tasks_reconciliation_data = process_Tasks_data(params, s3_bucket, local_file_name)
        ReconFilewriter.writerow([Tasks_reconciliation_data[0], Tasks_reconciliation_data[1], Tasks_reconciliation_data[2],
                                 Tasks_reconciliation_data[3], Tasks_reconciliation_data[4], Tasks_reconciliation_data[5]])

        # Delete source files as they are processed and copied to S3 additing timestamp
        logging.info(str(key_file_name) + ' will be deleted from S3 as ' + str(local_file_name) + ' is added to S3.')
        # Delete is not required for single file processing
        #s3_bucket.Object(key_file_name).delete()

    except botocore.exceptions.ClientError as error:
            if error.response['Error']['Code'] == "404":
                logging.info('================================================================================')                
                logging.info(str(key_file_name) + ' not available for processing.')
                logging.info('================================================================================')
                ReconFilewriter.writerow(['tasks', 0, 0, 0, 0, 0])
            else:
                logging.error(error)


    # Recon file close and move to S3 
    ReconFile.close()
    local_file_name = ReconFilename
    archive_file_name = 'archive/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
    logging.info(str(ReconFilename) + ' is transferred to S3 bucket.')

    # Log file close and move to S3
    logging.info(str(LogFilename) + ' will be closed and then transferred to S3 bucket.')
    logging.info('End of the Job')
    logging.shutdown()
    local_file_name = LogFilename
    archive_file_name = 'archive/' + local_file_name
    s3_bucket.upload_file(local_file_name, archive_file_name)
